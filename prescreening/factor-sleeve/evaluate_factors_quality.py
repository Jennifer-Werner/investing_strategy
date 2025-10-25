import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==============================
# CONFIGURATION
# ==============================
INPUT_FILE = "../output/sp1500_financials_filled.xlsx"
OUTPUT_FILE = "../output/sp1500_factor_quality_evaluated.xlsx"

WEIGHTS = {
    "value": 0.4,
    "profitability": 0.3,
    "quality": 0.3
}

# ==============================
# HELPER FUNCTIONS
# ==============================
def safe_div(a, b):
    if isinstance(a, (pd.Series, np.ndarray)) or isinstance(b, (pd.Series, np.ndarray)):
        a_arr = np.array(a)
        b_arr = np.array(b)
        if a_arr.ndim > 1:
            a_arr = a_arr[:, 0]
        if b_arr.ndim > 1:
            b_arr = b_arr[:, 0]
        result = np.where((b_arr == 0) | pd.isna(a_arr) | pd.isna(b_arr), np.nan, a_arr / b_arr)
        return pd.Series(result, index=a.index if isinstance(a, pd.Series) else None)
    else:
        if pd.isna(a) or pd.isna(b) or b == 0:
            return np.nan
        return a / b

def zscore(series):
    return stats.zscore(series, nan_policy='omit')

# ==============================
# LOAD DATA
# ==============================
df = pd.read_excel(INPUT_FILE)
df.columns = df.columns.str.strip()

if "Symbol" not in df.columns:
    raise KeyError("Missing 'Symbol' column.")
df["Symbol"] = df["Symbol"].astype(str).str.upper().str.strip()

if "EV" not in df.columns:
    ev_candidates = [col for col in df.columns if col.strip().lower() == "ev"]
    if ev_candidates:
        df["EV"] = df[ev_candidates[0]]
        print(f"Using '{ev_candidates[0]}' as EV column.")

# ==============================
# STANDARDIZE COLUMN NAMES
# ==============================
df.columns = df.columns.str.strip().str.lower()

# ==============================
# FCF SETUP
# ==============================
if "fcf" not in df.columns:
    fcf_aliases = ["fcf", "free cash flow", "freecashflow", "free cashflow", "fcf (a)", "fcf (ttm)", "fcf ttm", "freecashflow (ttm)"]
    found_fcf = None
    for alias in fcf_aliases:
        for col in df.columns:
            if col.strip().lower() == alias:
                df["fcf"] = df[col]
                found_fcf = col
                break
        if found_fcf:
            break
    if not found_fcf:
        ocf_col = None
        capex_col = None
        for col in df.columns:
            lcol = col.strip().lower()
            if lcol in ["operating cash flow", "ocf"]:
                ocf_col = col
            if lcol in ["capital expenditures", "capex"]:
                capex_col = col
        if ocf_col and capex_col:
            df["fcf"] = df[ocf_col] + df[capex_col]
        else:
            df["fcf"] = np.nan

if "eps ttm" not in df.columns and "earnings per share" in df.columns:
    df["eps ttm"] = df["earnings per share"]

# ==============================
# FACTOR CONSTRUCTION
# ==============================
df["ev/ebitda"] = safe_div(df["ev"], df["ebitda"])
df["p/b"] = safe_div(df["ev"], df["equity"])
df["earnings yield"] = safe_div(df["eps ttm"], df["ev"])

# Compute Long-Term Debt / Equity
if "long term debt" in df.columns and "equity" in df.columns:
    df["lt_d/e"] = safe_div(df["long term debt"], df["equity"])
    print("✅ Created 'lt_d/e' column as long term debt divided by equity.")
else:
    df["lt_d/e"] = np.nan
    print("⚠️ Could not compute 'lt_d/e'.")

# ==============================
# PREPARE OUTPUT
# ==============================
df["value score"] = np.nan
df["profitability score"] = np.nan
df["quality score"] = np.nan
df["composite score"] = np.nan

cols_to_keep = [
    "symbol",
    "value score", "profitability score", "quality score", "composite score",
    "ev/ebitda", "p/b", "earnings yield",
    "roa", "roic", "operating margin",
    "roe", "lt_d/e", "eps ttm"
]

df_to_save = df[cols_to_keep].copy()
df_to_save.to_excel(OUTPUT_FILE, index=False)

# ==============================
# ADD EXCEL FORMULAS
# ==============================
wb = load_workbook(OUTPUT_FILE)
ws = wb.active
nrows = ws.max_row
header = [cell.value for cell in ws[1]]
col_idx = {name: idx+1 for idx, name in enumerate(header)}

def zscore_formula(col_letter, row, start_row=2, end_row=nrows):
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    return f"IF(ISNUMBER({col_letter}{row}),({col_letter}{row}-AVERAGE({rng}))/STDEV.P({rng}),\"\")"

val_cols = ["ev/ebitda", "p/b", "earnings yield"]
prof_cols = ["roa", "roic", "operating margin"]
qual_cols = ["roe", "lt_d/e", "eps ttm"]

for row in range(2, nrows+1):
    val_z = [zscore_formula(get_column_letter(col_idx[col]), row) for col in val_cols]
    prof_z = [zscore_formula(get_column_letter(col_idx[col]), row) for col in prof_cols]
    qual_z = []
    for col in qual_cols:
        zf = zscore_formula(get_column_letter(col_idx[col]), row)
        if col == "lt_d/e":
            zf = f"-({zf})"
        qual_z.append(zf)

    ws.cell(row=row, column=col_idx["value score"]).value = f"=AVERAGE({','.join(val_z)})"
    ws.cell(row=row, column=col_idx["profitability score"]).value = f"=AVERAGE({','.join(prof_z)})"
    ws.cell(row=row, column=col_idx["quality score"]).value = f"=AVERAGE({','.join(qual_z)})"

    val_cell = ws.cell(row=row, column=col_idx["value score"]).coordinate
    prof_cell = ws.cell(row=row, column=col_idx["profitability score"]).coordinate
    qual_cell = ws.cell(row=row, column=col_idx["quality score"]).coordinate
    comp_formula = f"=0.4*{val_cell}+0.3*{prof_cell}+0.3*{qual_cell}"
    ws.cell(row=row, column=col_idx["composite score"]).value = comp_formula

if hasattr(wb, "calculation_properties"):
    wb.calculation_properties.fullCalcOnLoad = True
elif hasattr(wb, "calc_properties"):
    wb.calc_properties.fullCalcOnLoad = True

wb.save(OUTPUT_FILE)

print(f"✅ Factor evaluation saved to: {OUTPUT_FILE}")
print(f"✅ Top 10 stocks by composite score:\n", df.sort_values("composite score", ascending=False).head(10)[['symbol', 'composite score']])